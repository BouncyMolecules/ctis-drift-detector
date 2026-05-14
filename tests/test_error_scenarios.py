"""Resilience and error-scenario tests for CTIS Drift Detector (Streamlit stack + domain services).

These tests avoid a browser or bound Streamlit server. They invoke the same Python
callbacks the UI relies on (`CtisPublicClient`, storage, Plotly helpers, workbook/PDF
exporters, and narrow UI paths with stubbed ``streamlit`` globals) while external
HTTPS is exercised via deterministic ``httpx.MockTransport``.

**Operational meaning of “does not crash”**

- Happy-degradation scenarios (cold database, zero-row charts/exports, empty styling)
  finish without leaking exceptions across the guarded boundary asserted in each test.
- Failure-mode scenarios validate **typed**, auditable exceptions or operator-visible
  error copy instead of ambiguous interpreter faults.

**Resilience ledger**

Executing this file with ``python tests/test_error_scenarios.py`` registers a scoped
pytest plugin that prints a scored ASCII table in ``pytest_terminal_summary``.
Plain ``pytest …`` on the same tests still runs all cases but omits the ledger unless
you instantiate the plugin yourself.
"""

from __future__ import annotations

import json
import textwrap
from collections.abc import Callable
from io import BytesIO
from pathlib import Path
from typing import Any, Final
from unittest import mock

import httpx
import pandas as pd
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlmodel import col, select

from ctis_drift.core.ctis_api import (
    CTISAPIClient,
    CtisHttpResponseError,
    CtisMalformedJsonError,
    CtisPublicApiError,
    CtisPublicClient,
    CtisRateLimitError,
    CtisTimeoutError,
    CtisTransportError,
    TrialFullRecord,
)
from ctis_drift.core.storage import DriftRunRecord, StorageService, TrialSnapshotRecord
from ctis_drift.main import (
    _SS_REFRESH_TOKEN,
    _bump_data_refresh,
    _friendly_api_error,
    _init_session_defaults,
    _resolve_audit_history_sort_column,
    apply_risk_styling,
    build_monitored_dataframe,
    export_pdf_bytes,
    export_workbook_bytes,
    fetch_trial_records,
    fig_change_frequency_histogram,
    fig_risk_trend,
    latest_run_lookup,
    render_global_exports,
    run_ctis_check_and_persist,
)

# ---------------------------------------------------------------------------
# Ledger rows: (scenario_display_name, expected_outcome_statement, pytest_tail)
# pytest_tail MUST match pytest's node suffix: nodeid.split("::", 1)[-1],
# including parametrised brackets such as ``testFoo[retrieve_404]``.
# ---------------------------------------------------------------------------

_SCENARIO_ROWS: Final[list[tuple[str, str, str]]] = [
    (
        "API - empty EU CT rejected before HTTP",
        "ValueError before any outbound call",
        "test_api_empty_euct_rejected_before_http",
    ),
    (
        "API - whitespace-only EU CT rejected",
        "ValueError; rejects blank identifiers like the ingest field would normalize",
        "test_api_whitespace_only_euct_rejected",
    ),
    (
        "API - HTTP 404 retrieve (trial missing)",
        "CtisHttpResponseError with diagnostic status code",
        "test_api_http_errors[retrieve_404]",
    ),
    (
        "API - HTTP 500 retrieve (operator-facing server fault)",
        "CtisHttpResponseError surfaced after bounded retry semantics",
        "test_api_http_errors[retrieve_500_final]",
    ),
    (
        "API - network / TLS transport failure",
        "CtisTransportError captures endpoint context",
        "test_api_network_connection_error_mapped_to_transport",
    ),
    (
        "API - synchronous timeout exhaustion",
        "CtisTimeoutError after retry budget exhaustion",
        "test_api_timeout_mapped_to_timeout_error",
    ),
    (
        "API - HTTP 429 rate limit exhaustion",
        "CtisRateLimitError when CTIS mandates backoff repeatedly",
        "test_api_rate_limit_429_exhaustion",
    ),
    (
        "API - syntactically invalid JSON payload on success HTTP",
        "CtisMalformedJsonError (decode-safe guardrail)",
        "test_api_malformed_json_response",
    ),
    (
        "API - `{}` envelope fails TrialFullRecord schema (empty trial body)",
        "CtisMalformedJsonError from schema validation parity with UI explorer",
        "test_api_empty_object_retrieve_fails_schema",
    ),
    (
        "Storage - first-time SQLite cold start",
        "Registry + dataframe builders tolerate zero rows without fault",
        "test_storage_first_time_empty_database",
    ),
    (
        "Storage - empty snapshot lineage for unseen EU CT number",
        "History query returns deterministic empty sequence",
        "test_storage_empty_history_for_trial",
    ),
    (
        "Storage - drift export sort column absent `created_utc`",
        "Resolver gracefully falls through until it finds temporal columns",
        "test_storage_history_dataframe_missing_created_utc_column",
    ),
    (
        "Storage - corrupted snapshot blob in SQLite WAL",
        "JSONDecodeError propagates loudly for corrective DB remediation",
        "test_storage_corrupted_snapshot_json_raises_decode_error",
    ),
    (
        "Input policy - enrolling duplicate EU CT (hash-equal payload)",
        "Second ingest elides redundant immutable snapshot rows",
        "test_input_duplicate_trial_snapshot_idempotent",
    ),
    (
        "Input fuzz - elongated / Unicode-rich EU CT path segments",
        "Malformed IDs yield HTTP-tier errors rather than interpreter faults",
        "test_input_long_and_special_character_euct_paths",
    ),
    (
        "Validated envelope - disallow `{}` manual retrieve JSON",
        "Pydantic ValidationError aligns with guarded Streamlit ingestion path",
        "test_ui_empty_euct_ingest_payload_guard",
    ),
    (
        "Visualisation - activity histogram absent drift audits",
        "Plotly scaffold renders explicit empty-state copy",
        "test_viz_activity_histogram_empty_runs",
    ),
    (
        "Visualisation - risk trajectory with zero evaluations",
        "Figure factory stays callable; subplot titles explain missing history",
        "test_viz_risk_trend_empty_runs",
    ),
    (
        "Visualisation - single-point risk spline unavailable",
        "Marker chart path documents deferred spline trend",
        "test_viz_risk_trend_single_run_insufficient_for_spline_trend",
    ),
    (
        "Export - Excel workbook with empty monitoring + audit registers",
        "openpyxl round trip preserves expected sheet names",
        "test_export_excel_with_empty_frames",
    ),
    (
        "Export - PDF appendix with zero audit rows",
        "fpdf2 output remains a well-formed %PDF stream",
        "test_export_pdf_with_no_sample_rows",
    ),
    (
        "Workflow isolation - neighbouring trials survive one failed CTIS retrieve",
        "Sequential UX contract: BAD trial raises typed error; GOOD trial persists run",
        "test_robustness_one_trial_fails_others_can_succeed",
    ),
    (
        "Copy deck - `_friendly_api_error` transport branch",
        "Operator sees connectivity language without raw stack or exception strings",
        "test_robustness_friendly_api_error_copy_for_transport",
    ),
    (
        "Portfolio styling - dataframe heatmap skips empty grids",
        "apply_risk_styling short-circuit without throwing",
        "test_robustness_apply_risk_styling_empty_dataframe",
    ),
    (
        "Orchestration - `run_ctis_check_and_persist` leaky API faults",
        "CtisPublicApiError lineage bubbles to Streamlit catcher",
        "test_robustness_run_ctis_check_surfaces_api_errors",
    ),
    (
        "Session scaffold - deterministic defaults + optimistic refresh bumps",
        "Session tokens initialise + increment under mocked widgets",
        "test_session_state_init_and_refresh_bump_mocked_streamlit",
    ),
    (
        "Graceful degradation - global exports sidebar on empty datastore",
        "render_global_exports issues guidance without raising",
        "test_graceful_degradation_sidebar_exports_render_on_empty_database",
    ),
]


class ResilienceReportPlugin:
    """Session plugin: tracks call outcomes then prints ledger + score."""

    def __init__(self, ledger: list[tuple[str, str, str]]) -> None:
        self._ledger = ledger
        self._reports: dict[str, pytest.TestReport] = {}

    def pytest_runtest_logreport(self, report: pytest.TestReport) -> None:
        if report.when != "call":
            return
        key = report.nodeid.split("::", 1)[-1]
        self._reports[key] = report

    def pytest_terminal_summary(self, terminalreporter: pytest.TerminalReporter) -> None:
        terminalreporter.write_line("")
        terminalreporter.write_line(_format_resilience_table(self._ledger, self._reports))
        terminalreporter.write_line("")


def _format_resilience_table(
    rows: list[tuple[str, str, str]],
    reports: dict[str, pytest.TestReport],
) -> str:
    lines: list[str] = []
    lines.append("=" * 110)
    lines.append("CTIS Drift Detector - Resilience / error-scenario ledger")
    lines.append("=" * 110)
    header = f"{'Scenario':<46} | {'Result':<8} | {'Expected behaviour (abridged)':<44}"
    lines.append(header)
    lines.append("-" * 110)

    passed = failed = skipped = missing = 0
    for name, expected, tail in rows:
        rep = reports.get(tail)
        if rep is None:
            display = "Missing"
            missing += 1
        elif rep.outcome == "skipped":
            display = "Skipped"
            skipped += 1
        elif rep.outcome == "failed":
            display = "Failed"
            failed += 1
        else:
            display = "Passed"
            passed += 1
        abridged = textwrap.shorten(expected, width=44, placeholder=" …")
        lines.append(f"{name:<46} | {display:<8} | {abridged}")

    lines.append("-" * 110)
    total = len(rows)
    score = round(100.0 * (passed + 0.5 * skipped) / max(total, 1), 1)
    lines.append(
        f"Totals - Passed: {passed} | Failed: {failed} | Skipped: {skipped} | "
        f"Unmatched rows: {missing} | Declared scenarios: {total}",
    )
    lines.append(
        f"Overall resilience score: {score}% "
        "(pass=100%, skip=50%, fail/missing=0%; informational only).",
    )
    if missing:
        lines.append(
            "Hint: ledger keys follow pytest node suffixes; update _SCENARIO_ROWS "
            "if you rename tests or param ids.",
        )
    lines.append("=" * 110)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# HTTP lab harness
# ---------------------------------------------------------------------------


def _lab_client(
    handler: Callable[[httpx.Request], httpx.Response],
    *,
    max_retries: int = 1,
) -> CtisPublicClient:
    transport = httpx.MockTransport(handler)
    http_client = httpx.Client(transport=transport, timeout=httpx.Timeout(5.0))
    return CtisPublicClient(
        client=http_client,
        max_total_retries=max_retries,
        min_request_interval_seconds=0.0,
    )


def _trial_json(ct: str = "2024-518143-38-99", status: str = "Authorised") -> dict[str, Any]:
    return {"ctNumber": ct, "ctStatus": status}


# ---------------------------------------------------------------------------
# API layer
# ---------------------------------------------------------------------------


def test_api_empty_euct_rejected_before_http() -> None:
    """Blank EU CT numbers must error before httpx issues a request."""

    client = _lab_client(lambda _: httpx.Response(200, json=_trial_json()))
    with pytest.raises(ValueError, match="non-empty"):
        client.get_full_trial("")
    client.close()


def test_api_whitespace_only_euct_rejected() -> None:
    """Whitespace-only inputs normalise to empty and must be rejected."""

    client = _lab_client(lambda _: httpx.Response(200, json=_trial_json()))
    with pytest.raises(ValueError, match="non-empty"):
        client.get_full_trial("  \t  ")
    client.close()


@pytest.mark.parametrize(
    "case",
    ["retrieve_404", "retrieve_500_final"],
)
def test_api_http_errors(case: str) -> None:
    """HTTP error classes surface consistently for missing vs server-side faults."""

    if case == "retrieve_404":

        def handler(_request: httpx.Request) -> httpx.Response:
            return httpx.Response(404, json={"reason": "missing"})

    else:

        def handler(_request: httpx.Request) -> httpx.Response:
            return httpx.Response(500, json={"reason": "boom"})

    client = _lab_client(handler, max_retries=1)
    with pytest.raises(CtisHttpResponseError) as excinfo:
        client.get_full_trial("2024-000001-42-01")
    assert excinfo.value.status_code in {404, 500}
    client.close()


def test_api_network_connection_error_mapped_to_transport() -> None:
    """Low-level socket/TLS issues become CtisTransportError records."""

    def handler(request: httpx.Request) -> httpx.Response:
        raise httpx.ConnectError("simulated routing failure", request=request)

    client = _lab_client(handler, max_retries=1)
    with pytest.raises(CtisTransportError):
        client.get_full_trial("2024-000001-42-02")
    client.close()


def test_api_timeout_mapped_to_timeout_error() -> None:
    """TimeoutException must convert to CtisTimeoutError for UI messaging."""

    def handler(request: httpx.Request) -> httpx.Response:
        raise httpx.TimeoutException("simulated socket stall", request=request)

    client = _lab_client(handler, max_retries=1)
    with pytest.raises(CtisTimeoutError):
        client.get_full_trial("2024-000001-42-03")
    client.close()


def test_api_rate_limit_429_exhaustion() -> None:
    """429 responses should honor Retry-After hints until retries exhaust."""

    attempts = {"count": 0}

    def handler(_request: httpx.Request) -> httpx.Response:
        attempts["count"] += 1
        return httpx.Response(429, json={"status": "slow_down"}, headers={"Retry-After": "0"})

    client = _lab_client(handler, max_retries=2)
    with pytest.raises(CtisRateLimitError):
        client.get_full_trial("2024-000001-42-04")
    assert attempts["count"] >= 2
    client.close()


def test_api_malformed_json_response() -> None:
    """Non-JSON payloads on HTTP 200 must never deserialize silently."""

    client = _lab_client(lambda _req: httpx.Response(200, text="NOT JSON {{{"), max_retries=1)
    with pytest.raises(CtisMalformedJsonError):
        client.get_full_trial("2024-000001-42-05")
    client.close()


def test_api_empty_object_retrieve_fails_schema() -> None:
    """Empty JSON object fails TrialFullRecord validation → MalformedJson family."""

    client = _lab_client(lambda _req: httpx.Response(200, json={}), max_retries=1)
    with pytest.raises(CtisMalformedJsonError):
        client.get_full_trial("2024-000001-42-06")
    client.close()


# ---------------------------------------------------------------------------
# Storage & projection layer
# ---------------------------------------------------------------------------


def test_storage_first_time_empty_database(tmp_path: Path) -> None:
    """Fresh SQLite file: zero trials, safe dataframe assembly."""

    db_url = f"sqlite:///{(tmp_path / 'cold_start.db').as_posix()}"
    storage = StorageService(db_url)
    storage.init_db()
    trials = fetch_trial_records(storage)
    assert trials == []
    df = build_monitored_dataframe(trials, lookup=latest_run_lookup(storage.recent_runs(limit=50)))
    assert df.empty


def test_storage_empty_history_for_trial(tmp_path: Path) -> None:
    """No snapshots yet → chronological history is an empty list."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'history_empty.db').as_posix()}")
    storage.init_db()
    assert storage.get_history("2024-never-seen-00-01", limit=5) == []


def test_storage_history_dataframe_missing_created_utc_column() -> None:
    """Export sort helper must not assume `created_utc` always exists."""

    frame = pd.DataFrame(
        {"timestamp": pd.to_datetime(["2021-01-01", "2022-01-01"], utc=True)},
    )
    assert _resolve_audit_history_sort_column(frame) == "timestamp"


def test_storage_corrupted_snapshot_json_raises_decode_error(tmp_path: Path) -> None:
    """Tampered snapshot rows should fail JSON decoding loudly for DB ops."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'corrupt_blob.db').as_posix()}")
    storage.init_db()
    euct = "2024-corrupt-00-01"
    storage.save_snapshot(euct, _trial_json(ct=euct))
    with storage.session() as session:
        row = session.exec(
            select(TrialSnapshotRecord).where(col(TrialSnapshotRecord.euct_number) == euct),
        ).first()
        assert row is not None
        row.raw_json = "{not-json"
        session.add(row)
        session.commit()
    with pytest.raises(json.JSONDecodeError):
        storage.get_latest_snapshot(euct)


# ---------------------------------------------------------------------------
# Input / governance semantics
# ---------------------------------------------------------------------------


def test_input_duplicate_trial_snapshot_idempotent(tmp_path: Path) -> None:
    """Identical CTIS JSON must not spam duplicate immutable rows."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'dedupe.db').as_posix()}")
    storage.init_db()
    euct = "2024-dup-42-07"
    payload = _trial_json(ct=euct)
    first = storage.save_snapshot(euct, payload, skip_duplicate_hash=True)
    assert first.persisted_new_row is True
    second = storage.save_snapshot(euct, payload, skip_duplicate_hash=True)
    assert second.persisted_new_row is False


def test_input_long_and_special_character_euct_paths() -> None:
    """Hostile strings should still route through httpx without Python faults."""

    long_id = "2024-" + ("x" * 80) + "-00-01"
    special = "2024-weird /\u0394-!?-99"

    def handler(_request: httpx.Request) -> httpx.Response:
        return httpx.Response(404, json={"detail": "unknown"})

    client = _lab_client(handler, max_retries=1)
    for euct in (long_id, special):
        with pytest.raises(CtisHttpResponseError) as exc:
            client.get_full_trial(euct)
        assert exc.value.status_code == 404
    client.close()


def test_ui_empty_euct_ingest_payload_guard() -> None:
    """Mirrors strict pydantic guard used once JSON returns from CTIS."""

    with pytest.raises(ValidationError):
        TrialFullRecord.model_validate({})


# ---------------------------------------------------------------------------
# Plotly visualisation
# ---------------------------------------------------------------------------


def test_viz_activity_histogram_empty_runs() -> None:
    """Zero-length drift runs must still produce a titled figure."""

    fig = fig_change_frequency_histogram([])
    title = fig.layout.title
    text = title.text if hasattr(title, "text") else str(title)
    assert "drift" in text.lower() or "activity" in text.lower()


def test_viz_risk_trend_empty_runs() -> None:
    """Risk splines require points; empty input should degrade gracefully."""

    fig = fig_risk_trend([], title="Empty trial history")
    title = fig.layout.title
    text = title.text if hasattr(title, "text") else str(title)
    assert "no data" in text.lower()


def test_viz_risk_trend_single_run_insufficient_for_spline_trend() -> None:
    """Single evaluation uses marker path + explicit subtitle about splines."""

    run = DriftRunRecord(
        trial_id="2024-single-42-09",
        metric_name="trial_snapshot",
        drift_score=0.42,
        method="pytest_stub",
        details_json=json.dumps({"risk_level": "HIGH", "risk_score": 0.72}),
    )
    fig = fig_risk_trend([run], title="Under-sampled timeline")
    title = fig.layout.title
    subtitle = getattr(title, "subtitle", None)
    subtitle_txt = getattr(subtitle, "text", "") if subtitle is not None else ""
    assert "single" in subtitle_txt.lower() or "evaluation" in subtitle_txt.lower()


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------


def test_export_excel_with_empty_frames() -> None:
    """Workbook writer must tolerate TMF-style empty registers."""

    monitored = pd.DataFrame(
        columns=[
            "EU CT Number",
            "Last polled (UTC)",
            "Content fingerprint",
            "Last drift check (UTC)",
            "Metric",
            "Risk score",
            "Risk band",
        ],
    )
    hist = pd.DataFrame(
        columns=[
            "created_utc",
            "trial_id",
            "metric_name",
            "drift_score_0_1",
            "method",
            "risk_band",
        ],
    )
    raw = export_workbook_bytes(monitored, hist)
    wb = load_workbook(BytesIO(raw))
    assert "Monitoring_register" in wb.sheetnames
    assert "Drift_runs_audit" in wb.sheetnames


def test_export_pdf_with_no_sample_rows() -> None:
    """PDF exporter should emit header/footer even when audit rows absent."""

    pdf = export_pdf_bytes("0 trials registered — cold start snapshot.", [])
    assert pdf.startswith(b"%PDF")


# ---------------------------------------------------------------------------
# Robustness & session behaviour
# ---------------------------------------------------------------------------


def test_robustness_one_trial_fails_others_can_succeed(tmp_path: Path) -> None:
    """Ensures per-trial fetch failures do not poison subsequent successes."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'parallel.db').as_posix()}")
    storage.init_db()

    good_euct = "2024-good-42-10"
    good_record = TrialFullRecord.model_validate(_trial_json(ct=good_euct))

    api_bad = mock.Mock(spec=CTISAPIClient)
    api_bad.get_full_trial.side_effect = CtisTransportError(
        "simulated outage",
        url="https://euclinicaltrials.eu/ctis-public-api/retrieve/BAD",
    )

    with pytest.raises(CtisTransportError):
        run_ctis_check_and_persist(
            storage=storage,
            euct="BAD-2024-518143",
            api_client=api_bad,
            persist_snapshot=True,
        )

    api_ok = mock.Mock(spec=CTISAPIClient)
    api_ok.get_full_trial.return_value = good_record

    report = run_ctis_check_and_persist(
        storage=storage,
        euct=good_euct,
        api_client=api_ok,
        persist_snapshot=True,
    )
    assert report.trial_id == good_euct


def test_robustness_friendly_api_error_copy_for_transport() -> None:
    """Transport branch must stay aligned with Streamlit `st.error` copy."""

    exc = CtisTransportError("forced transport", url="https://example/ctis")
    message = _friendly_api_error(exc)
    assert "could not reach" in message.lower()


def test_robustness_apply_risk_styling_empty_dataframe() -> None:
    """Heatmap styling should no-op on empty portfolio tables."""

    empty = pd.DataFrame()
    styled = apply_risk_styling(empty)
    frame = getattr(styled, "_data", styled)
    assert isinstance(frame, pd.DataFrame)
    assert frame.empty


def test_robustness_run_ctis_check_surfaces_api_errors(tmp_path: Path) -> None:
    """Typed API errors bubble to UI catchers instead of generic crashes."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'api_fail.db').as_posix()}")
    storage.init_db()
    api = mock.Mock(spec=CTISAPIClient)
    api.get_full_trial.side_effect = CtisHttpResponseError(
        "HTTP 502",
        url="https://euclinicaltrials.eu/ctis-public-api/retrieve/X",
        status_code=502,
        body_preview="gateway",
    )
    with pytest.raises(CtisPublicApiError):
        run_ctis_check_and_persist(
            storage=storage,
            euct="2024-fail-42-11",
            api_client=api,
            persist_snapshot=False,
        )


def test_session_state_init_and_refresh_bump_mocked_streamlit() -> None:
    """Session defaults + refresh bump must work without a hosted Streamlit kernel."""

    fake_state: dict[str, Any] = {}
    with mock.patch("ctis_drift.ui.streamlit_env.st") as st_mock:
        st_mock.session_state = fake_state
        _init_session_defaults()
        assert _SS_REFRESH_TOKEN in fake_state
        _bump_data_refresh()
        assert fake_state[_SS_REFRESH_TOKEN] == 1
        _bump_data_refresh()
        assert fake_state[_SS_REFRESH_TOKEN] == 2


def test_graceful_degradation_sidebar_exports_render_on_empty_database(tmp_path: Path) -> None:
    """Sidebar export stack should warn/inform instead of aborting on cold DB."""

    storage = StorageService(f"sqlite:///{(tmp_path / 'exports_only.db').as_posix()}")
    storage.init_db()

    def _column_stub() -> Any:
        cm = mock.MagicMock()
        cm.__enter__.return_value = cm
        cm.__exit__.return_value = False
        return cm

    col_left = _column_stub()
    col_right = _column_stub()

    with mock.patch("ctis_drift.ui.streamlit_env.st") as st_mock:
        st_mock.session_state = {}
        st_mock.info = mock.Mock()
        st_mock.caption = mock.Mock()
        st_mock.columns = mock.Mock(return_value=(col_left, col_right))
        st_mock.error = mock.Mock()
        st_mock.download_button = mock.Mock()
        render_global_exports(storage)
        st_mock.info.assert_called()


if __name__ == "__main__":
    plugin = ResilienceReportPlugin(_SCENARIO_ROWS)
    raise SystemExit(
        pytest.main(
            [
                str(Path(__file__).resolve()),
                "-v",
                "--tb=short",
            ],
            plugins=[plugin],
        ),
    )
