"""Full-stack regression tests against the live CTIS public API.

These tests always call the public read-only endpoints used by the dashboard
(``euclinicaltrials.eu``): search, full trial retrieve, then local storage, drift
evaluation, and export helpers. Requests are paced (~0.22 s minimum interval) with
retries to stay polite on the shared production rate limits.

**Safety / prerequisites**

- Outbound HTTPS must be allowed; failures are environmental, not application skips.
- Trial payloads are read-only; one scenario mutates an *in-memory* copy after fetch
  to assert drift detection (nothing is written back to CTIS).

**Environment variables**

``CTIS_DRIFT_E2E_TRIALS`` (optional)
    Comma-separated EU CT numbers that exist in the public registry. When set,
    at least three identifiers are required; up to five are used. When unset,
    three distinct trial IDs are taken from the first page of the public
    ``/search`` response.

Run only this module::

    pytest tests/test_end_to_end.py -v

Filter by marker (same tests; markers describe behaviour)::

    pytest tests/test_end_to_end.py -m "e2e and network" -v
"""

from __future__ import annotations

import copy
import json
import os
from collections.abc import Generator
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from ctis_drift.core.ctis_api import (
    CtisPublicApiError,
    CtisPublicClient,
    SearchPagination,
    TrialFullRecord,
    TrialSearchPayload,
    get_full_trial,
    search_trials,
)
from ctis_drift.core.drift_detector import (
    DriftDetector,
    RiskLevel,
    build_regulatory_report,
    risk_level_from_score,
    snapshot_content_hash,
)
from ctis_drift.core.storage import (
    DriftRunRecord,
    StorageService,
    compute_json_sha256,
    normalize_json_bytes,
)
from ctis_drift.main import (
    build_monitored_dataframe,
    export_pdf_bytes,
    export_workbook_bytes,
    fetch_trial_records,
    latest_run_lookup,
)

pytestmark = [
    pytest.mark.e2e,
    pytest.mark.network,
]


@pytest.fixture(scope="module")
def ctis_public_client() -> Generator[CtisPublicClient, None, None]:
    """Shared, courteously paced HTTP client (module scope amortises TLS handshakes)."""

    with CtisPublicClient(
        min_request_interval_seconds=0.22,
        timeout_seconds=90.0,
        max_total_retries=5,
    ) as client:
        yield client


@pytest.fixture(scope="module")
def live_euct_numbers(ctis_public_client: CtisPublicClient) -> tuple[str, ...]:
    """Resolve at least three public EU CT numbers — env override or search discovery."""

    override = os.environ.get("CTIS_DRIFT_E2E_TRIALS", "").strip()
    if override:
        ids = tuple(x.strip() for x in override.split(",") if x.strip())
        if len(ids) < 3:
            pytest.fail(
                "CTIS_DRIFT_E2E_TRIALS must list at least three comma-separated EU CT numbers.",
            )
        return ids[:5]

    response = ctis_public_client.search_trials(
        TrialSearchPayload(pagination=SearchPagination(page=1, size=12))
    )
    numbers = tuple({row.ct_number for row in response.data})
    if len(numbers) < 3:
        pytest.fail(
            "CTIS search returned fewer than three distinct EU CT numbers; "
            "try again later or populate CTIS_DRIFT_E2E_TRIALS.",
        )
    return numbers[:3]


@pytest.fixture
def isolated_storage(tmp_path: Path) -> StorageService:
    db_path = tmp_path / "e2e_isolated.db"
    service = StorageService(f"sqlite:///{db_path.as_posix()}")
    service.init_db()
    return service


def _payload_from_full_trial_record(record: TrialFullRecord) -> dict[str, Any]:
    dumped = record.model_dump_json(by_alias=True)
    parsed = json.loads(dumped)
    if not isinstance(parsed, dict):
        msg = "CTIS retrieve payload must deserialize to a JSON object"
        raise TypeError(msg)
    return dict(parsed)


def test_ctis_api_search_and_retrieve_round_trip(
    ctis_public_client: CtisPublicClient,
    live_euct_numbers: tuple[str, ...],
) -> None:
    """High-level helpers and the typed client agree on envelopes for search + retrieve."""

    search_via_module = search_trials(
        payload=TrialSearchPayload(pagination=SearchPagination(page=1, size=5)),
        use_streamlit_cache=False,
    )
    search_via_client = ctis_public_client.search_trials(
        TrialSearchPayload(pagination=SearchPagination(page=1, size=5))
    )
    assert search_via_module.pagination.total_records == search_via_client.pagination.total_records
    assert {row.ct_number for row in search_via_module.data} == {
        row.ct_number for row in search_via_client.data
    }

    target = live_euct_numbers[0]
    record_cached_disabled = get_full_trial(target, use_streamlit_cache=False)
    record_client = ctis_public_client.get_full_trial(target)

    assert record_cached_disabled.ct_number == record_client.ct_number
    assert record_cached_disabled.ct_number == target

    payload_a = _payload_from_full_trial_record(record_cached_disabled)
    payload_b = _payload_from_full_trial_record(record_client)
    assert compute_json_sha256(payload_a) == compute_json_sha256(payload_b)


@pytest.mark.parametrize("euct_index", (0, 1, 2))
def test_full_happy_path_storage_detector_report_persist_and_exports(
    ctis_public_client: CtisPublicClient,
    isolated_storage: StorageService,
    live_euct_numbers: tuple[str, ...],
    euct_index: int,
) -> None:
    """Mirrors production ordering: retrieve → drift vs storage → snapshot → audit row → exports."""

    euct = live_euct_numbers[euct_index]
    record = ctis_public_client.get_full_trial(euct)
    raw = _payload_from_full_trial_record(record)

    assert isolated_storage.has_changed(euct, raw).changed is True

    detector = DriftDetector()
    pre_anchor_report = detector.evaluate_with_storage(isolated_storage, euct, raw)
    assert pre_anchor_report.trial_id == euct
    assert pre_anchor_report.changed is True

    assert "No stored baseline snapshot." in (
        pre_anchor_report.detailed_diff.structural.ingest_notes or ""
    )

    assert 0 <= pre_anchor_report.risk_score <= 100
    assert pre_anchor_report.risk_level == risk_level_from_score(pre_anchor_report.risk_score)
    assert euct in pre_anchor_report.human_readable_summary
    assert len(pre_anchor_report.human_readable_summary) >= 48

    anchor = isolated_storage.save_snapshot(euct, raw, skip_duplicate_hash=True)
    assert anchor.persisted_new_row is True
    assert anchor.content_hash == compute_json_sha256(raw)
    duplicate_attempt = isolated_storage.save_snapshot(euct, raw, skip_duplicate_hash=True)
    assert duplicate_attempt.persisted_new_row is False

    assert isolated_storage.has_changed(euct, raw).changed is False

    steady_report = detector.evaluate_with_storage(isolated_storage, euct, raw)
    assert steady_report.changed is False
    assert steady_report.risk_score == 0
    assert steady_report.risk_level is RiskLevel.LOW
    assert "no cryptographic drift" in steady_report.human_readable_summary.lower()

    loaded = isolated_storage.get_latest_snapshot(euct)
    assert loaded is not None
    assert compute_json_sha256(raw) == compute_json_sha256(loaded)
    canon_raw = json.loads(normalize_json_bytes(raw).decode("utf-8"))
    canon_loaded = json.loads(normalize_json_bytes(loaded).decode("utf-8"))
    assert canon_raw == canon_loaded

    for rep in (pre_anchor_report, steady_report):
        row = isolated_storage.save_report(rep)
        assert row.trial_id == rep.trial_id
        assert row.metric_name == rep.metric_name
        assert abs(float(row.drift_score) - rep.drift_score) < 1e-9
        blob = json.loads(row.details_json or "{}")
        assert isinstance(blob, dict)
        assert int(blob.get("risk_score", -1)) == rep.risk_score

    runs = isolated_storage.recent_runs(limit=50)
    lookup = latest_run_lookup(runs)
    trials = fetch_trial_records(isolated_storage)
    monitored = build_monitored_dataframe(trials, lookup=lookup)
    assert euct in set(monitored["EU CT Number"].astype(str))

    xbytes = export_workbook_bytes(monitored, pd_audit_frame(runs))
    wb = load_workbook(BytesIO(xbytes))
    assert "Monitoring_register" in wb.sheetnames
    assert "Drift_runs_audit" in wb.sheetnames
    assert wb["Monitoring_register"].max_row >= 1

    pdf_bytes = export_pdf_bytes(
        f"E2E export probe - {euct} - {len(runs)} drift rows.",
        [
            {
                "trial_id": euct,
                "created_utc": "1970-01-01T00:00:00+00:00",
                "metric_name": "trial_snapshot",
                "drift_score_0_1": "0.0000",
                "method": "regulatory_snapshot_v1",
            },
        ],
    )
    assert pdf_bytes.startswith(b"%PDF")


def test_simulated_regulatory_change_after_baseline(
    ctis_public_client: CtisPublicClient,
    isolated_storage: StorageService,
    live_euct_numbers: tuple[str, ...],
) -> None:
    """Loads a real baseline, perturbs regulator-visible fields, and re-evaluates."""

    euct = live_euct_numbers[0]
    baseline_payload = _payload_from_full_trial_record(ctis_public_client.get_full_trial(euct))
    isolated_storage.save_snapshot(euct, baseline_payload, skip_duplicate_hash=True)

    baseline_hash = compute_json_sha256(baseline_payload)
    meta = isolated_storage.get_latest_snapshot_record(euct)
    assert meta is not None
    assert meta.data_hash == baseline_hash

    detector = DriftDetector()
    steady = detector.evaluate_with_storage(isolated_storage, euct, baseline_payload)
    assert steady.changed is False

    mutated = copy.deepcopy(baseline_payload)
    mutated_original_status = mutated.get("ctStatus")
    mutated["ctStatus"] = f"SIMULATED_E2E_SHIFT_FROM_{mutated_original_status}"
    mutated.setdefault("trialRfis", [])
    if isinstance(mutated["trialRfis"], list):
        mutated["trialRfis"] = list(mutated["trialRfis"]) + [
            {
                "e2eSyntheticRfi": True,
                "topic": "regulatory-clarification",
                "note": "Synthetic row for drift detector ontology checks",
            },
        ]

    change = isolated_storage.has_changed(euct, mutated)
    assert change.changed is True
    assert change.summary.previous_hash == baseline_hash

    report_eval = detector.evaluate_with_storage(isolated_storage, euct, mutated)
    assert report_eval.changed is True
    assert report_eval.risk_score > 0
    assert report_eval.risk_level == risk_level_from_score(report_eval.risk_score)
    assert "Detected substantive drift" in report_eval.human_readable_summary
    assert euct in report_eval.human_readable_summary

    direct = build_regulatory_report(
        trial_id=euct,
        baseline_payload=baseline_payload,
        candidate_payload=mutated,
        previous_content_hash=baseline_hash,
        current_content_hash=snapshot_content_hash(mutated),
        ingest_notes=change.summary.notes,
        baseline_timestamp=meta.timestamp,
        changed_explicit=True,
    )
    assert direct.changed is True
    assert direct.risk_score == report_eval.risk_score
    assert direct.risk_level == report_eval.risk_level

    isolated_storage.save_report(report_eval)
    isolated_storage.save_snapshot(euct, mutated, skip_duplicate_hash=True)
    reloaded = isolated_storage.get_latest_snapshot(euct)
    assert reloaded is not None
    assert compute_json_sha256(mutated) == compute_json_sha256(reloaded)


def test_public_api_errors_surface_consistently(ctis_public_client: CtisPublicClient) -> None:
    """Malformed identifiers should fail fast with the project's API error types."""

    with pytest.raises((CtisPublicApiError, ValueError)):
        ctis_public_client.get_full_trial("")


def pd_audit_frame(runs: list[DriftRunRecord]) -> Any:
    """Build the same drift-history frame shape used by the Streamlit export path."""

    import pandas as pd

    return pd.DataFrame(
        [
            {
                "created_utc": r.created_at.isoformat(timespec="seconds"),
                "trial_id": r.trial_id,
                "metric_name": r.metric_name,
                "drift_score_0_1": r.drift_score,
                "method": r.method,
            }
            for r in runs
        ],
    )
